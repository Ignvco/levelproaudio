# apps/products/importers.py

import openpyxl
import csv
import io
from decimal import Decimal, InvalidOperation
from django.utils.text import slugify
from .models import Product, Category, Brand, ProductImage


REQUIRED_COLUMNS = ["nombre", "precio", "stock"]

COLUMN_MAP = {
    # nombre_excel          : campo_modelo
    "nombre":               "name",
    "descripcion":          "description",
    "descripcion_corta":    "short_description",
    "precio":               "price",
    "precio_comparacion":   "compare_price",
    "stock":                "stock",
    "sku":                  "sku",
    "categoria":            "category",
    "marca":                "brand",
    "peso":                 "weight",
    "destacado":            "is_featured",
    "activo":               "is_active",
    "tipo":                 "product_type",
    "imagen_url":           "image_url",
    "seo_titulo":           "seo_title",
    "seo_descripcion":      "seo_description",
}


def normalize_header(h: str) -> str:
    """Normaliza encabezados: quita espacios, tildes, minúsculas."""
    return (
        h.strip().lower()
        .replace("á","a").replace("é","e").replace("í","i")
        .replace("ó","o").replace("ú","u").replace("ñ","n")
        .replace(" ", "_")
    )


def parse_bool(val) -> bool:
    if isinstance(val, bool): return val
    return str(val).strip().lower() in ("si", "sí", "true", "1", "yes", "x")


def parse_decimal(val) -> Decimal:
    try:
        cleaned = str(val).replace("$","").replace(".","").replace(",",".").strip()
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def get_or_create_category(name: str) -> Category | None:
    if not name or str(name).strip() == "":
        return None
    name = str(name).strip()
    # Soporte para subcategorías: "Audio Pro > Mezcladores"
    if ">" in name:
        parts = [p.strip() for p in name.split(">")]
        parent = None
        for part in parts:
            cat, _ = Category.objects.get_or_create(
                slug=slugify(part),
                defaults={"name": part, "parent": parent}
            )
            parent = cat
        return parent
    cat, _ = Category.objects.get_or_create(
        slug=slugify(name),
        defaults={"name": name}
    )
    return cat


def get_or_create_brand(name: str) -> Brand | None:
    if not name or str(name).strip() == "":
        return None
    name = str(name).strip()
    brand, _ = Brand.objects.get_or_create(
        slug=slugify(name),
        defaults={"name": name}
    )
    return brand


def parse_rows(rows: list[dict]) -> dict:
    """
    Procesa filas del excel/csv y retorna:
    {
        "created": [...],
        "updated": [...],
        "errors":  [{"row": N, "name": ..., "error": ...}],
        "skipped": 0,
    }
    """
    created = []
    updated = []
    errors  = []

    for i, row in enumerate(rows, start=2):  # start=2 porque fila 1 es header
        name = str(row.get("name", "")).strip()
        if not name:
            errors.append({"row": i, "name": "—", "error": "Nombre vacío, fila ignorada."})
            continue

        try:
            price = parse_decimal(row.get("price", 0))
            if price <= 0:
                errors.append({"row": i, "name": name, "error": "Precio inválido o 0."})
                continue

            stock = int(float(str(row.get("stock", 0)).replace(",", ".")))

            # Slug único
            base_slug = slugify(name)
            slug      = base_slug
            counter   = 1
            while Product.objects.filter(slug=slug).exclude(
                name__iexact=name
            ).exists():
                slug = f"{base_slug}-{counter}"
                counter += 1

            category  = get_or_create_category(row.get("category", ""))
            brand     = get_or_create_brand(row.get("brand", ""))

            compare_price = parse_decimal(row.get("compare_price", 0)) or None
            weight        = parse_decimal(row.get("weight", 0)) or None
            is_featured   = parse_bool(row.get("is_featured", False))
            is_active     = parse_bool(row.get("is_active", True))
            product_type  = str(row.get("product_type", "physical")).strip() or "physical"

            product, was_created = Product.objects.update_or_create(
                name__iexact=name,
                defaults={
                    "name":              name,
                    "slug":              slug,
                    "description":       str(row.get("description", "")),
                    "short_description": str(row.get("short_description", "")),
                    "price":             price,
                    "compare_price":     compare_price,
                    "stock":             stock,
                    "sku":               str(row.get("sku", "")).strip() or None,
                    "category":          category,
                    "brand":             brand,
                    "weight":            weight,
                    "is_featured":       is_featured,
                    "is_active":         is_active,
                    "product_type":      product_type,
                    "seo_title":         str(row.get("seo_title", ""))[:70],
                    "seo_description":   str(row.get("seo_description", ""))[:160],
                },
            )

            # Imagen principal por URL
            image_url = str(row.get("image_url", "")).strip()
            if image_url and was_created:
                ProductImage.objects.get_or_create(
                    product=product,
                    defaults={
                        "image_url": image_url,
                        "is_primary": True,
                        "order": 0,
                    }
                )

            if was_created:
                created.append(name)
            else:
                updated.append(name)

        except Exception as e:
            errors.append({"row": i, "name": name, "error": str(e)})

    return {
        "created": created,
        "updated": updated,
        "errors":  errors,
        "total_processed": len(created) + len(updated),
    }


def import_from_excel(file) -> dict:
    wb   = openpyxl.load_workbook(file, data_only=True)
    ws   = wb.active

    headers_raw = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers     = [normalize_header(h) for h in headers_raw]
    mapped      = [COLUMN_MAP.get(h, h) for h in headers]

    # Valida columnas requeridas
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        return {"error": f"Faltan columnas requeridas: {', '.join(missing)}"}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(mapped, row)))

    return parse_rows(rows)


def import_from_csv(file) -> dict:
    content  = file.read().decode("utf-8-sig")
    reader   = csv.DictReader(io.StringIO(content))
    headers  = [normalize_header(h) for h in (reader.fieldnames or [])]

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        return {"error": f"Faltan columnas requeridas: {', '.join(missing)}"}

    rows = []
    for row in reader:
        mapped_row = {COLUMN_MAP.get(normalize_header(k), normalize_header(k)): v
                      for k, v in row.items()}
        rows.append(mapped_row)

    return parse_rows(rows)