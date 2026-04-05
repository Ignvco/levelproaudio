# apps/analytics/views.py — VERSION LIMPIA SIN DUPLICADOS

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework import serializers as drf_serializers
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Sum, Count as DCount, Avg, Q, F, ExpressionWrapper, DecimalField
from django.db.models import Count
from django.db.models import F as Fdb
from django.db.models.functions import TruncDay, TruncMonth
from django.utils import timezone as tz
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from datetime import timedelta
from core.admin_permissions import IsAdminOrStaff

from apps.orders.models import Order, OrderItem, OrderStatus
from apps.orders.serializers import OrderSerializer
from apps.products.models import Product, Category, Brand, ProductImage
from apps.products.serializers import (
    ProductDetailSerializer, ProductListSerializer, ProductImageSerializer, CategorySerializer, BrandSerializer
)
import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.utils.text import slugify

from decimal import Decimal
from apps.users.models import User
from apps.payments.models import Payment
from apps.academy.models import Course, Module, Lesson, Enrollment
from apps.services.models import Service, Booking, ServiceRequest, ServiceCategory
from apps.products.importers import import_from_excel, import_from_csv
from apps.analytics.models import DistributionCategory, IncomeRecord, IncomeDistribution, Withdrawal
from apps.services.serializers import ServiceDetailSerializer

# ═══════════════════════════════════════════════════════════
# SERIALIZERS
# ═══════════════════════════════════════════════════════════

class AdminCategorySerializer(drf_serializers.ModelSerializer):
    class Meta:
        model  = Category
        fields = ["id", "name", "slug", "description",
                  "parent", "is_active", "order", "image"]
        extra_kwargs = {
            "slug": {"required": False, "allow_blank": True},
        }


class AdminBrandSerializer(drf_serializers.ModelSerializer):
    class Meta:
        model  = Brand
        fields = ["id", "name", "slug", "website", "is_active", "logo"]
        extra_kwargs = {
            "slug": {"required": False, "allow_blank": True},
        }


class AdminModuleSerializer(drf_serializers.ModelSerializer):
    course_title = drf_serializers.CharField(source="course.title", read_only=True)

    class Meta:
        model  = Module
        fields = ["id", "title", "order", "course", "course_title"]


class AdminLessonSerializer(drf_serializers.ModelSerializer):
    module_title = drf_serializers.CharField(source="module.title", read_only=True)

    class Meta:
        model  = Lesson
        fields = ["id", "title", "video_url", "description",
                  "order", "is_free", "duration_minutes", "module", "module_title"]


class CourseAdminSerializer(drf_serializers.ModelSerializer):
    enrollment_count = drf_serializers.SerializerMethodField()
    total_lessons    = drf_serializers.SerializerMethodField()
    thumbnail_url    = drf_serializers.SerializerMethodField()

    class Meta:
        model  = Course
        fields = [
            "id", "title", "slug", "short_description", "description",
            "thumbnail", "thumbnail_url", "preview_url",
            "price", "level", "is_free", "is_published", "order",
            "enrollment_count", "total_lessons",
        ]
        extra_kwargs = {
            "slug": {"required": False, "allow_blank": True},
        }

    def get_thumbnail_url(self, obj):
        # Path relativo — proxy Vite lo resuelve en dev
        if obj.thumbnail:
            return obj.thumbnail.url
        return None

    def get_enrollment_count(self, obj):
        return getattr(obj, "enrollment_count", obj.enrollments.count())

    def get_total_lessons(self, obj):
        return obj.total_lessons


class AdminProductWriteSerializer(drf_serializers.ModelSerializer):
    """Serializer para crear/editar productos — acepta IDs de categoría y marca."""
    class Meta:
        model  = Product
        fields = [
            "id", "name", "slug", "sku",
            "category", "brand",
            "short_description", "description",
            "price", "compare_price", "stock",
            "product_type", "is_active", "is_featured",
            "weight", "seo_title", "seo_description",
        ]
        extra_kwargs = {
            "slug": {"required": False, "allow_blank": True},
        }

        

# ═══════════════════════════════════════════════════════════
# VIEWSETS
# ═══════════════════════════════════════════════════════════

class AdminOrderViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    serializer_class   = OrderSerializer
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ["email", "id"]
    ordering_fields    = ["created_at", "total", "status"]
    ordering           = ["-created_at"]

    def get_queryset(self):
        qs = Order.objects.select_related("user").prefetch_related(
            "items", "payments"
        ).all()
        status_filter = self.request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs

    @action(detail=True, methods=["post"])
    def update_status(self, request, pk=None):
        from apps.analytics.models import IncomeRecord
        from apps.payments.models import Payment
        from decimal import Decimal
        from django.utils import timezone

        order      = self.get_object()
        new_status = request.data.get("status")
        valid      = [s[0] for s in OrderStatus.choices]

        if new_status not in valid:
            return Response({"error": "Estado inválido."}, status=400)

        order.status = new_status
        order.save(update_fields=["status"])

        # ── Sincroniza pagos ─────────────────────────────────────
        pago = order.payments.first()

        if new_status in ["paid", "shipped", "completed"]:
            # Aprueba el pago si existía pendiente
            if pago and pago.status != "approved":
                pago.status  = "approved"
                pago.paid_at = timezone.now()
                pago.save(update_fields=["status", "paid_at"])

            # ── Crea IncomeRecord si no existe ───────────────────
            if not IncomeRecord.objects.filter(order=order).exists():
                record = IncomeRecord.objects.create(
                    order       = order,
                    payment     = pago,
                    amount      = Decimal(str(order.total)),
                    description = f"Venta — Orden #{str(order.id)[:8].upper()}",
                    source      = "payment",
                )
                record.create_distributions()

        elif new_status == "cancelled":
            # Cancela el pago
            if pago and pago.status not in ["refunded"]:
                pago.status = "cancelled"
                pago.save(update_fields=["status"])

            # Elimina IncomeRecord
            IncomeRecord.objects.filter(order=order).delete()

        return Response({"status": new_status, "detail": "Estado actualizado."})


class AdminProductViewSet(viewsets.ModelViewSet):
    """CRUD completo de productos para el admin."""
    permission_classes = [IsAdminOrStaff]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ["name", "sku"]
    ordering_fields    = ["name", "price", "stock", "created_at"]
    ordering           = ["-created_at"]
    lookup_field       = "id"

    def get_queryset(self):
        return Product.objects.select_related(
            "category", "brand"
        ).prefetch_related("images").all()

    def get_serializer_class(self):
        # Lectura → serializer completo con relaciones
        if self.action in ["list", "retrieve"]:
            return ProductDetailSerializer
        # Escritura → procesamos manualmente para aceptar IDs y cost_price
        return ProductDetailSerializer  # fallback

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context

    def _save_product(self, request, product=None):
        from decimal import Decimal, InvalidOperation
        from django.utils.text import slugify

        data = request.data

        if product is None:
            product = Product()

        # ── Campos de texto ──────────────────────────────────────
        text_fields = [
            "name", "sku", "short_description", "description",
            "product_type", "seo_title", "seo_description",
        ]
        for field in text_fields:
            if field in data:
                val = data[field]
                # SKU vacío → None (respeta UNIQUE)
                if field == "sku" and val == "":
                    val = None
                setattr(product, field, val)

        # ── Campos decimales — siempre convertir a Decimal ───────
        decimal_fields = ["price", "compare_price", "cost_price"]
        for field in decimal_fields:
            if field in data:
                val = data[field]
                if val == "" or val is None:
                    setattr(product, field, None)
                else:
                    try:
                        setattr(product, field, Decimal(str(val)))
                    except (InvalidOperation, TypeError):
                        setattr(product, field, None)

        # ── Stock — entero ───────────────────────────────────────
        if "stock" in data:
            try:
                product.stock = int(data["stock"])
            except (ValueError, TypeError):
                product.stock = 0

        # ── Booleanos ────────────────────────────────────────────
        bool_fields = ["is_active", "is_featured"]
        for field in bool_fields:
            if field in data:
                val = data[field]
                if isinstance(val, str):
                    val = val.lower() in ["true", "1", "yes"]
                setattr(product, field, val)

        # ── Slug automático ──────────────────────────────────────
        if not product.slug and product.name:
            product.slug = slugify(product.name)

        # ── Relaciones por ID ────────────────────────────────────
        if "category" in data:
            cat_id = data["category"]
            product.category_id = cat_id if cat_id else None

        if "brand" in data:
            brand_id = data["brand"]
            product.brand_id = brand_id if brand_id else None

        product.save()
        return product

    
    def create(self, request, *args, **kwargs):
        try:
            product = self._save_product(request)
            serializer = ProductDetailSerializer(
                product, context=self.get_serializer_context()
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        product = self.get_object()
        try:
            product = self._save_product(request, product)
            serializer = ProductDetailSerializer(
                product, context=self.get_serializer_context()
            )
            return Response(serializer.data)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)


class AdminCategoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = Category.objects.all().order_by("order", "name")
    filter_backends    = [filters.SearchFilter]
    search_fields      = ["name"]

    # ← usa siempre AdminCategorySerializer (no lo pisa con get_serializer_class)
    serializer_class   = AdminCategorySerializer
    
    def get_serializer_class(self):
        return CategorySerializer

    def perform_create(self, serializer):
        serializer.save()

    def perform_update(self, serializer):
        serializer.save()



class AdminBrandViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    serializer_class   = AdminBrandSerializer
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    queryset           = Brand.objects.all().order_by("name")
    filter_backends    = [filters.SearchFilter]
    search_fields      = ["name"]
    
    def get_serializer_class(self):
        return BrandSerializer

    def perform_create(self, serializer):
        serializer.save()


class AdminProductImageViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    parser_classes     = [MultiPartParser, FormParser]
    serializer_class   = ProductImageSerializer

    def get_queryset(self):
        product_pk = self.kwargs.get("product_pk")
        if product_pk:
            return ProductImage.objects.filter(
                product_id=product_pk
            ).order_by("order")
        return ProductImage.objects.all().order_by("order")

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context
    
    def create(self, request, *args, **kwargs):
        product_pk = self.kwargs.get("product_pk")
        if not product_pk:
            return Response(
                {"error": "product_pk requerido"},
                status=status.HTTP_400_BAD_REQUEST
            )

        image_file = request.FILES.get("image")
        if not image_file:
            return Response(
                {"error": "No se recibió ningún archivo. Envía 'image' como campo."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            product = Product.objects.get(id=product_pk)
        except Product.DoesNotExist:
            return Response({"error": "Producto no encontrado."}, status=404)

        is_first = not ProductImage.objects.filter(product=product).exists()

        img = ProductImage.objects.create(
            product    = product,
            image      = image_file,
            alt_text   = request.data.get("alt_text", product.name),
            order      = int(request.data.get("order", 0)),
            is_primary = (
                request.data.get("is_primary", "false").lower() == "true"
                or is_first
            ),
        )

        serializer = ProductImageSerializer(img, context={"request": request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def perform_create(self, serializer):
        from apps.products.models import Product
        product = Product.objects.get(id=self.kwargs["product_pk"])
        serializer.save(product=product)


class AdminCourseViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    serializer_class   = CourseAdminSerializer
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    lookup_field       = "id"

    def get_queryset(self):
        return Course.objects.annotate(
            enrollment_count=Count("enrollments")
        ).order_by("-created_at")


class AdminModuleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    serializer_class   = AdminModuleSerializer
    parser_classes     = [JSONParser]

    def get_queryset(self):
        qs = Module.objects.all().order_by("order")
        course_id = self.request.query_params.get("course")
        if course_id:
            qs = qs.filter(course_id=course_id)
        return qs


class AdminLessonViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAdminOrStaff]
    serializer_class   = AdminLessonSerializer
    parser_classes     = [JSONParser]

    def get_queryset(self):
        qs = Lesson.objects.all().order_by("order")
        module_id = self.request.query_params.get("module")
        if module_id:
            qs = qs.filter(module_id=module_id)
        return qs


# ═══════════════════════════════════════════════════════════
# API VIEWS (funciones)
# ═══════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def dashboard_overview(request):
    now         = timezone.now()
    today       = now.date()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_30     = now - timedelta(days=30)
    last_7      = now - timedelta(days=7)

    paid_orders = Order.objects.filter(status__in=["paid", "shipped", "completed"])

    total_revenue      = paid_orders.aggregate(t=Sum("total"))["t"] or 0
    revenue_this_month = paid_orders.filter(
        created_at__gte=month_start
    ).aggregate(t=Sum("total"))["t"] or 0
    revenue_last_30    = paid_orders.filter(
        created_at__gte=last_30
    ).aggregate(t=Sum("total"))["t"] or 0

    total_orders      = Order.objects.count()
    pending_orders    = Order.objects.filter(status="pending").count()
    orders_today      = Order.objects.filter(created_at__date=today).count()
    orders_this_month = Order.objects.filter(created_at__gte=month_start).count()

    total_users     = User.objects.count()
    new_users_month = User.objects.filter(date_joined__gte=month_start).count()
    new_users_week  = User.objects.filter(date_joined__gte=last_7).count()

    total_products = Product.objects.filter(is_active=True).count()
    low_stock      = Product.objects.filter(is_active=True, stock__lte=5).count()
    out_of_stock   = Product.objects.filter(is_active=True, stock=0).count()

    total_enrollments = Enrollment.objects.count()
    enrollments_month = Enrollment.objects.filter(
        created_at__gte=month_start
    ).count()

    sales_chart = (
        paid_orders
        .filter(created_at__gte=last_30)
        .annotate(day=TruncDay("created_at"))
        .values("day")
        .annotate(total=Sum("total"), count=Count("id"))
        .order_by("day")
    )

    top_products = (
        OrderItem.objects
        .filter(order__status__in=["paid", "shipped", "completed"])
        .values("product_name")
        .annotate(total_sold=Sum("quantity"), revenue=Sum("price"))
        .order_by("-total_sold")[:5]
    )

    recent_orders = Order.objects.select_related("user").order_by("-created_at")[:8]

    return Response({
        "revenue": {
            "total":        float(total_revenue),
            "this_month":   float(revenue_this_month),
            "last_30_days": float(revenue_last_30),
        },
        "orders": {
            "total":      total_orders,
            "pending":    pending_orders,
            "today":      orders_today,
            "this_month": orders_this_month,
        },
        "users": {
            "total":     total_users,
            "new_month": new_users_month,
            "new_week":  new_users_week,
        },
        "products": {
            "total":        total_products,
            "low_stock":    low_stock,
            "out_of_stock": out_of_stock,
        },
        "academy": {
            "total_enrollments": total_enrollments,
            "enrollments_month": enrollments_month,
        },
        "sales_chart": [
            {
                "date":  item["day"].strftime("%Y-%m-%d"),
                "total": float(item["total"]),
                "count": item["count"],
            }
            for item in sales_chart
        ],
        "top_products": list(top_products),
        "recent_orders": [
            {
                "id":      str(o.id),
                "email":   o.email,
                "total":   float(o.total),
                "status":  o.status,
                "created": o.created_at.strftime("%Y-%m-%d %H:%M"),
            }
            for o in recent_orders
        ],
    })


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_users(request):
    search = request.query_params.get("search", "")
    users  = User.objects.all().order_by("-date_joined")

    if search:
        users = users.filter(
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )

    data = []
    for u in users[:100]:
        orders_count = Order.objects.filter(user=u).count()
        total_spent  = Order.objects.filter(
            user=u, status__in=["paid", "shipped", "completed"]
        ).aggregate(t=Sum("total"))["t"] or 0
        data.append({
            "id":           str(u.id),
            "email":        u.email,
            "first_name":   u.first_name,
            "last_name":    u.last_name,
            "date_joined":  u.date_joined.strftime("%Y-%m-%d"),
            "is_active":    u.is_active,
            "orders_count": orders_count,
            "total_spent":  float(total_spent),
        })

    return Response({"results": data, "count": users.count()})

@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdminOrStaff])
def admin_user_detail(request, user_id):
    """PATCH/DELETE /api/v1/admin/users/{id}/"""
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"error": "Usuario no encontrado."}, status=404)

    if request.method == "PATCH":
        allowed = ["first_name", "last_name", "phone", "is_active", "is_staff"]
        for field in allowed:
            if field in request.data:
                setattr(user, field, request.data[field])
        user.save(update_fields=[f for f in allowed if f in request.data])
        return Response({"detail": "Usuario actualizado."})

    if request.method == "DELETE":
        user.delete()
        return Response({"detail": "Usuario eliminado."}, status=204)


@api_view(["DELETE"])
@permission_classes([IsAdminOrStaff])
def admin_delete_enrollment(request, enrollment_id):
    """DELETE /api/v1/admin/academy/enrollments/{id}/"""
    try:
        enrollment = Enrollment.objects.get(id=enrollment_id)
        enrollment.delete()
        return Response({"deleted": True})
    except Enrollment.DoesNotExist:
        return Response({"error": "No encontrada."}, status=404)

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_payments(request):
    status_filter = request.query_params.get("status")
    payments = Payment.objects.select_related("order").order_by("-created_at")

    if status_filter:
        payments = payments.filter(status=status_filter)

    data = [{
        "id":          str(p.id),
        "order_id":    str(p.order.id),
        "order_email": p.order.email,
        "provider":    p.provider,
        "status":      p.status,
        "amount":      float(p.amount),
        "currency":    p.currency,
        "paid_at":     p.paid_at.strftime("%Y-%m-%d %H:%M") if p.paid_at else None,
        "created_at":  p.created_at.strftime("%Y-%m-%d %H:%M"),
    } for p in payments[:200]]

    total_approved = payments.filter(
        status="approved"
    ).aggregate(t=Sum("amount"))["t"] or 0

    return Response({"results": data, "total_approved": float(total_approved)})

@api_view(["PATCH"])
@permission_classes([IsAdminOrStaff])
def admin_update_payment_status(request, payment_id):
    """PATCH /api/v1/admin/payments/{id}/status/"""
    

    try:
        payment = Payment.objects.select_related("order").get(id=payment_id)
    except Payment.DoesNotExist:
        return Response({"error": "Pago no encontrado."}, status=404)

    new_status = request.data.get("status")
    valid = ["pending", "approved", "rejected", "cancelled", "refunded"]
    if new_status not in valid:
        return Response({"error": f"Estado inválido. Opciones: {valid}"}, status=400)

    old_status = payment.status
    payment.status = new_status

    if new_status == "approved" and old_status != "approved":
        from django.utils import timezone
        payment.paid_at = timezone.now()

    payment.save()

    # ── Sincroniza la orden ──────────────────────────────────
    order = payment.order
    if new_status == "approved":
        order.status = "paid"
        order.save(update_fields=["status"])

        # ── Crea IncomeRecord si no existe ───────────────────
        if not IncomeRecord.objects.filter(order=order).exists():
            record = IncomeRecord.objects.create(
                payment     = payment,
                order       = order,
                amount      = Decimal(str(payment.amount)),
                description = f"Pago aprobado — Orden #{str(order.id)[:8].upper()}",
                source      = "payment",
            )
            record.create_distributions()

    elif new_status in ["rejected", "cancelled"]:
        # Cancela la orden si no hay otro pago aprobado
        other_approved = order.payments.filter(
            status="approved"
        ).exclude(id=payment.id).exists()
        if not other_approved:
            order.status = "cancelled"
            order.save(update_fields=["status"])
            # Elimina IncomeRecord si existía
            IncomeRecord.objects.filter(order=order).delete()

    elif new_status == "refunded":
        order.status = "cancelled"
        order.save(update_fields=["status"])
        # Marca el IncomeRecord como reembolsado eliminándolo
        IncomeRecord.objects.filter(order=order).delete()

    return Response({
        "id":         str(payment.id),
        "status":     payment.status,
        "order_id":   str(order.id),
        "order_status": order.status,
    })
    
    

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_academy(request):
   
    courses = Course.objects.annotate(
        enrollment_count=DCount("enrollments", distinct=True),
        lessons_count=DCount("modules__lessons", distinct=True),  # ← agrega
    ).order_by("-created_at")

    return Response({
        "results": [{
            "id":               str(c.id),
            "title":            c.title,
            "slug":             c.slug,
            "price":            float(c.price),
            "is_published":     c.is_published,
            "is_free":          c.is_free,
            "level":            c.level,
            "enrollment_count": c.enrollment_count,
            "total_lessons":    c.lessons_count,  # ← usa anotación, no property
            "thumbnail":        c.thumbnail.url if c.thumbnail else None,
        } for c in courses]
    })


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_services(request):
    bookings = Booking.objects.select_related(
        "user", "service"
    ).order_by("-created_at")[:50]

    requests_qs = ServiceRequest.objects.select_related(
        "user", "service"
    ).order_by("-created_at")[:50]

    return Response({
        "bookings": [{
            "id":             str(b.id),
            "user_email":     b.user.email,
            "service_name":   b.service.name,
            "scheduled_date": b.scheduled_date.strftime("%Y-%m-%d %H:%M"),
            "status":         b.status,
            "notes":          b.notes,
        } for b in bookings],
        "requests": [{
            "id":           str(r.id),
            "name":         r.name,
            "email":        r.email,
            "service_name": r.service.name if r.service else "General",
            "message":      r.message,
            "budget":       float(r.budget) if r.budget else None,
            "status":       r.status,
            "created_at":   r.created_at.strftime("%Y-%m-%d %H:%M"),
        } for r in requests_qs],
    })

@api_view(["GET", "POST"])
@permission_classes([IsAdminOrStaff])
def admin_services_crud(request):
    """GET/POST /api/v1/admin/services/list/"""
   

    if request.method == "GET":
        services = Service.objects.select_related("category").order_by("order", "name")
        return Response({
            "results": [{
                "id":               str(s.id),
                "name":             s.name,
                "slug":             s.slug,
                "category_name":    s.category.name if s.category else "—",
                "price_display":    s.price_display,
                "price_type":       s.price_type,
                "is_active":        s.is_active,
                "is_featured":      s.is_featured,
                "thumbnail":        s.thumbnail.url if s.thumbnail else None,
            } for s in services]
        })

    # POST — crear servicio
    
    data = request.data
    service = Service.objects.create(
        name             = data.get("name", ""),
        slug             = slugify(data.get("name", "")),
        short_description= data.get("short_description", ""),
        description      = data.get("description", ""),
        price_type       = data.get("price_type", "quote"),
        price            = data.get("price") or None,
        is_active        = data.get("is_active", True),
        is_featured      = data.get("is_featured", False),
        order            = data.get("order", 0),
    )
    return Response({"id": str(service.id), "name": service.name}, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdminOrStaff])
def admin_service_detail(request, service_id):
    """PATCH/DELETE /api/v1/admin/services/{id}/"""
    
    try:
        service = Service.objects.get(id=service_id)
    except Service.DoesNotExist:
        return Response({"error": "No encontrado."}, status=404)

    if request.method == "DELETE":
        service.delete()
        return Response({"deleted": True})

    allowed = ["name", "short_description", "description", "price_type",
               "price", "is_active", "is_featured", "order"]
    for field in allowed:
        if field in request.data:
            setattr(service, field, request.data[field])
    service.save()
    return Response({"id": str(service.id), "name": service.name})

@api_view(["PATCH"])
@permission_classes([IsAdminOrStaff])
def update_booking_status(request, pk):
    try:
        booking = Booking.objects.get(id=pk)
    except Booking.DoesNotExist:
        return Response({"error": "No encontrado."}, status=404)
    booking.status = request.data.get("status", booking.status)
    booking.save(update_fields=["status"])
    return Response({"status": booking.status})


@api_view(["PATCH"])
@permission_classes([IsAdminOrStaff])
def update_request_status(request, pk):
    try:
        sr = ServiceRequest.objects.get(id=pk)
    except ServiceRequest.DoesNotExist:
        return Response({"error": "No encontrado."}, status=404)

    old_status     = sr.status
    sr.status      = request.data.get("status", sr.status)
    sr.admin_notes = request.data.get("admin_notes", sr.admin_notes)
    sr.save(update_fields=["status", "admin_notes"])

    # Acepta → crea Booking automáticamente
    if sr.status == "accepted" and old_status != "accepted" and sr.service:
        
        scheduled = (
            tz.datetime.combine(
                sr.preferred_date,
                tz.datetime.min.time()
            ).replace(tzinfo=tz.get_current_timezone())
            if sr.preferred_date
            else tz.now() + timedelta(days=7)
        )
        Booking.objects.get_or_create(
            user=sr.user,
            service=sr.service,
            defaults={
                "scheduled_date": scheduled,
                "notes":          sr.message,
                "status":         "confirmed",
            }
        )

    return Response({"status": sr.status})


@api_view(["POST"])
@permission_classes([IsAdminOrStaff])
def admin_create_enrollment(request):
    user_id   = request.data.get("user_id")
    course_id = request.data.get("course_id")
    try:
        user   = User.objects.get(id=user_id)
        course = Course.objects.get(id=course_id)
    except (User.DoesNotExist, Course.DoesNotExist) as e:
        return Response({"error": str(e)}, status=404)

    enrollment, created = Enrollment.objects.get_or_create(
        user=user, course=course
    )
    return Response({
        "created":       created,
        "enrollment_id": str(enrollment.id),
        "user":          user.email,
        "course":        course.title,
    })


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_enrollments(request):
    enrollments = Enrollment.objects.select_related(
        "user", "course"
    ).order_by("-created_at")[:200]

    return Response({
        "results": [{
            "id":           str(e.id),
            "user_email":   e.user.email,
            "user_name":    f"{e.user.first_name} {e.user.last_name}".strip(),
            "course_title": e.course.title,
            "course_id":    str(e.course.id),
            "progress":     e.progress_percentage,
            "created_at":   e.created_at.strftime("%Y-%m-%d"),
        } for e in enrollments]
    })


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def analytics_full(request):
    now      = timezone.now()
    last_12m = now - timedelta(days=365)
    last_30  = now - timedelta(days=30)

    paid_orders = Order.objects.filter(
        status__in=["paid", "shipped", "completed"]
    )

    sales_by_month = (
        paid_orders.filter(created_at__gte=last_12m)
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Sum("total"), count=Count("id"))
        .order_by("month")
    )

    orders_by_status = (
        Order.objects
        .values("status")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    users_by_month = (
        User.objects.filter(date_joined__gte=last_12m)
        .annotate(month=TruncMonth("date_joined"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    revenue_by_provider = (
        Payment.objects.filter(status="approved")
        .values("provider")
        .annotate(total=Sum("amount"), count=Count("id"))
        .order_by("-total")
    )

    low_stock_products = (
        Product.objects
        .filter(is_active=True, stock__lte=10)
        .order_by("stock")
        .values("name", "stock", "sku")[:15]
    )

    revenue_by_category = (
        OrderItem.objects
        .filter(
            order__status__in=["paid", "shipped", "completed"],
            product__isnull=False
        )
        .values("product__category__name")
        .annotate(total=Sum("price"), count=Sum("quantity"))
        .order_by("-total")[:8]
    )

    enrollments_by_month = (
        Enrollment.objects.filter(created_at__gte=last_12m)
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    total_revenue    = paid_orders.aggregate(t=Sum("total"))["t"] or 0
    avg_order_value  = paid_orders.aggregate(a=Avg("total"))["a"] or 0
    total_customers  = Order.objects.values("user").distinct().count()
    repeat_customers = (
        Order.objects.values("user")
        .annotate(c=Count("id"))
        .filter(c__gt=1)
        .count()
    )

    return Response({
        "kpis": {
            "total_revenue":    float(total_revenue),
            "avg_order_value":  float(avg_order_value),
            "total_customers":  total_customers,
            "repeat_customers": repeat_customers,
            "repeat_rate": round(
                (repeat_customers / total_customers * 100)
                if total_customers else 0, 1
            ),
        },
        "sales_by_month": [
            {"month": i["month"].strftime("%b %Y"),
             "total": float(i["total"]), "count": i["count"]}
            for i in sales_by_month
        ],
        "orders_by_status": [
            {"status": i["status"], "count": i["count"]}
            for i in orders_by_status
        ],
        "users_by_month": [
            {"month": i["month"].strftime("%b %Y"), "count": i["count"]}
            for i in users_by_month
        ],
        "revenue_by_provider": [
            {"provider": i["provider"],
             "total": float(i["total"]), "count": i["count"]}
            for i in revenue_by_provider
        ],
        "low_stock_products": list(low_stock_products),
        "revenue_by_category": [
            {"category": i["product__category__name"] or "Sin categoría",
             "total": float(i["total"]), "count": i["count"]}
            for i in revenue_by_category
        ],
        "enrollments_by_month": [
            {"month": i["month"].strftime("%b %Y"), "count": i["count"]}
            for i in enrollments_by_month
        ],
    })

@api_view(["POST"])
@permission_classes([IsAdminOrStaff])
def import_products(request):
    """
    POST /api/v1/admin/products/import/
    Acepta .xlsx o .csv y crea/actualiza productos masivamente.
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No se envió ningún archivo."}, status=400)

    name = file.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        result = import_from_excel(file)
    elif name.endswith(".csv"):
        result = import_from_csv(file)
    else:
        return Response({"error": "Formato no soportado. Usa .xlsx o .csv."}, status=400)

    if "error" in result:
        return Response(result, status=400)

    return Response(result, status=200)


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def download_template(request):
    """
    GET /api/v1/admin/products/template/
    Descarga el Excel plantilla con todas las columnas y ejemplos.
    """
    

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "nombre", "descripcion_corta", "descripcion", "precio",
        "precio_comparacion", "stock", "sku", "categoria",
        "marca", "peso", "destacado", "activo", "tipo",
        "imagen_url", "seo_titulo", "seo_descripcion",
    ]

    # Estilos encabezado
    header_font  = Font(bold=True, color="000000", size=11)
    header_fill  = PatternFill("solid", fgColor="1AFF6E")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = max(16, len(header) + 4)

    # Fila de ejemplo
    examples = [
        "Shure SM58",
        "Micrófono vocal dinámico profesional",
        "El Shure SM58 es el micrófono vocal más utilizado en escenarios del mundo.",
        "150000",
        "180000",
        "10",
        "SM58-LC",
        "Micrófonos",
        "Shure",
        "0.3",
        "Si",
        "Si",
        "physical",
        "https://ejemplo.com/imagen.jpg",
        "Shure SM58 - Micrófono Vocal Profesional",
        "Compra el Shure SM58, el micrófono más confiable del mercado.",
    ]

    example_font = Font(color="555555", size=10, italic=True)
    for col, val in enumerate(examples, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = example_font

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES DE USO", True),
        ("", False),
        ("COLUMNAS REQUERIDAS (obligatorias):", True),
        ("• nombre       → Nombre del producto (debe ser único)", False),
        ("• precio       → Precio en pesos, sin puntos ni $. Ej: 150000", False),
        ("• stock        → Cantidad disponible. Ej: 10", False),
        ("", False),
        ("COLUMNAS OPCIONALES:", True),
        ("• descripcion_corta   → Texto breve para listados", False),
        ("• descripcion         → Descripción completa del producto", False),
        ("• precio_comparacion  → Precio tachado (precio original antes del descuento)", False),
        ("• sku                 → Código interno del producto", False),
        ("• categoria           → Nombre de categoría. Para subcategoría: Audio Pro > Mezcladores", False),
        ("• marca               → Nombre de la marca", False),
        ("• peso                → Peso en kg. Ej: 0.5", False),
        ("• destacado           → Si / No (muestra en home)", False),
        ("• activo              → Si / No (visible en tienda)", False),
        ("• tipo                → physical / digital / course", False),
        ("• imagen_url          → URL de imagen principal", False),
        ("• seo_titulo          → Título para Google (máx 70 caracteres)", False),
        ("• seo_descripcion     → Descripción para Google (máx 160 caracteres)", False),
        ("", False),
        ("COMPORTAMIENTO:", True),
        ("• Si el producto ya existe (mismo nombre), se ACTUALIZA.", False),
        ("• Si no existe, se CREA.", False),
        ("• Las categorías y marcas se crean automáticamente si no existen.", False),
        ("• Puedes subir el archivo cuantas veces quieras, es seguro.", False),
    ]

    ws2.column_dimensions["A"].width = 70
    for row_num, (text, bold) in enumerate(instrucciones, 1):
        cell = ws2.cell(row=row_num, column=1, value=text)
        cell.font = Font(bold=bold, size=10 if not bold else 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_productos_levelproaudio.xlsx"'
    wb.save(response)
    return response

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def product_images(request, product_id):
    images = ProductImage.objects.filter(product_id=product_id).order_by("order")
    return Response([{
        "id":         str(img.id),
        "image":      img.image.url if img.image else None,  # ← relativa
        "alt_text":   img.alt_text,
        "order":      img.order,
        "is_primary": img.is_primary,
    } for img in images])


@api_view(["POST"])
@permission_classes([IsAdminOrStaff])
def product_image_upload(request, product_id):
    
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return Response({"error": "Producto no encontrado."}, status=404)

    image_file = request.FILES.get("image")
    if not image_file:
        return Response({"error": "No se recibió imagen."}, status=400)

    is_primary = not ProductImage.objects.filter(
        product=product, is_primary=True
    ).exists()

    img = ProductImage.objects.create(
        product    = product,
        image      = image_file,
        alt_text   = request.data.get("alt_text", product.name),
        is_primary = is_primary,
        order      = ProductImage.objects.filter(product=product).count(),
    )
    return Response({
        "id":         str(img.id),
        "image":      img.image.url if img.image else None,  # ← relativa
        "alt_text":   img.alt_text,
        "order":      img.order,
        "is_primary": img.is_primary,
    }, status=201)

@api_view(["DELETE"])
@permission_classes([IsAdminOrStaff])
def product_image_delete(request, image_id):
    """DELETE /api/v1/admin/images/{id}/delete/"""
    try:
        img = ProductImage.objects.get(id=image_id)
        img.image.delete(save=False)
        img.delete()
        return Response({"deleted": True})
    except ProductImage.DoesNotExist:
        return Response({"error": "No encontrada."}, status=404)


@api_view(["PATCH"])
@permission_classes([IsAdminOrStaff])
def product_image_set_primary(request, image_id):
    """PATCH /api/v1/admin/images/{id}/primary/"""
    try:
        img = ProductImage.objects.get(id=image_id)
        ProductImage.objects.filter(product=img.product).update(is_primary=False)
        img.is_primary = True
        img.save(update_fields=["is_primary"])
        return Response({"is_primary": True})
    except ProductImage.DoesNotExist:
        return Response({"error": "No encontrada."}, status=404)
    
    
    # ── MÓDULO FINANCIERO ────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAdminOrStaff])
def admin_finance_categories(request):
    """GET/POST /api/v1/admin/finance/categories/"""
    

    if request.method == "GET":
        cats = DistributionCategory.objects.all().order_by("order")
        return Response([{
            "id":         str(c.id),
            "name":       c.name,
            "cat_type":   c.cat_type,      
            "percentage": float(c.percentage),
            "color":      c.color,
            "icon":       c.icon,
            "order":      c.order,
            "is_active":  c.is_active,
        } for c in cats])

    # POST — nueva categoría
    cat = DistributionCategory.objects.create(
        name       = request.data.get("name", "Nueva categoría"),
        cat_type   = request.data.get("cat_type", "percent"),    # ← AGREGA
        percentage = float(request.data.get("percentage", 0)),   # ← float()
        color      = request.data.get("color", "#888888"),
        icon       = request.data.get("icon", "💰"),
        is_active  = request.data.get("is_active", True),        # ← AGREGA
        order      = DistributionCategory.objects.count() + 1,
    )
    return Response({"id": str(cat.id), "name": cat.name}, status=201)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdminOrStaff])
def admin_finance_category_detail(request, cat_id):
    """PATCH/DELETE /api/v1/admin/finance/categories/{id}/"""
    
    try:
        cat = DistributionCategory.objects.get(id=cat_id)
    except DistributionCategory.DoesNotExist:
        return Response({"error": "No encontrada."}, status=404)

    if request.method == "DELETE":
        cat.delete()
        return Response({"deleted": True})

    if request.method == "PATCH":
        for field in ["name", "cat_type", "color", "icon", "order", "is_active"]:  # ← cat_type
            if field in request.data:
                setattr(cat, field, request.data[field])

    if "percentage" in request.data:                    # ← float separado
        cat.percentage = float(request.data["percentage"])

    cat.save()

    # Recalcula distribuciones si cambiaron los %
    if "percentage" in request.data or "cat_type" in request.data:
        
        IncomeDistribution.objects.all().delete()
        for record in IncomeRecord.objects.all():
            record.create_distributions()

    return Response({
        "id":         str(cat.id),
        "name":       cat.name,
        "cat_type":   cat.cat_type,
        "percentage": float(cat.percentage),
        "color":      cat.color,
        "icon":       cat.icon,
        "is_active":  cat.is_active,
    })

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_finance_summary(request):
    """GET /api/v1/admin/finance/summary/"""
    

    categories = DistributionCategory.objects.filter(is_active=True)
    summary = []

    for cat in categories:
        total_in  = IncomeDistribution.objects.filter(
            category=cat
        ).aggregate(t=Sum("amount"))["t"] or 0

        total_out = Withdrawal.objects.filter(
            category=cat
        ).aggregate(t=Sum("amount"))["t"] or 0

        balance = float(total_in) - float(total_out)

        summary.append({
            "id":         str(cat.id),
            "name":       cat.name,
            "cat_type":   cat.cat_type,
            "percentage": float(cat.percentage),
            "color":      cat.color,
            "icon":       cat.icon,
            "total_in":   float(total_in),
            "total_out":  float(total_out),
            "balance":    balance,
        })

    # Totales generales
    total_income = IncomeRecord.objects.aggregate(
        t=Sum("amount"))["t"] or 0
    total_withdrawn = Withdrawal.objects.aggregate(
        t=Sum("amount"))["t"] or 0

    # Últimos ingresos
    
    recent = IncomeRecord.objects.select_related(
        "order", "payment"
    ).order_by("-created_at")[:20]

    recent_data = [{
        "id":          str(r.id),
        "amount":      float(r.amount),
        "description": r.description,
        "source":      r.source,
        "created_at":  r.created_at.strftime("%d/%m/%Y %H:%M"),
        "distributions": [{
            "category_name":  d.category.name,
            "category_color": d.category.color,
            "category_icon":  d.category.icon,
            "amount":         float(d.amount),
        } for d in r.distributions.all()],
    } for r in recent]

    return Response({
        "summary":          summary,
        "total_income":     float(total_income),
        "total_withdrawn":  float(total_withdrawn),
        "net_balance":      float(total_income) - float(total_withdrawn),
        "recent_income":    recent_data,
    })


@api_view(["GET", "POST"])
@permission_classes([IsAdminOrStaff])
def admin_finance_withdrawals(request):
    """GET/POST /api/v1/admin/finance/withdrawals/"""
    
    if request.method == "GET":
        ws = Withdrawal.objects.select_related("category").order_by("-created_at")[:50]
        return Response([{
            "id":            str(w.id),
            "category_name": w.category.name,
            "category_color":w.category.color,
            "category_icon": w.category.icon,
            "amount":        float(w.amount),
            "destination":   w.destination,
            "notes":         w.notes,
            "reference":     w.reference,
            "created_at":    w.created_at.strftime("%d/%m/%Y %H:%M"),
        } for w in ws])

    # POST — registrar retiro
    try:
        cat = DistributionCategory.objects.get(id=request.data.get("category_id"))
    except DistributionCategory.DoesNotExist:
        return Response({"error": "Categoría no encontrada."}, status=400)

    w = Withdrawal.objects.create(
        category    = cat,
        amount      = request.data.get("amount"),
        destination = request.data.get("destination", ""),
        notes       = request.data.get("notes", ""),
        reference   = request.data.get("reference", ""),
    )
    return Response({"id": str(w.id), "created": True}, status=201)


@api_view(["DELETE"])
@permission_classes([IsAdminOrStaff])
def admin_finance_withdrawal_delete(request, withdrawal_id):
    """DELETE /api/v1/admin/finance/withdrawals/{id}/"""
    
    try:
        w = Withdrawal.objects.get(id=withdrawal_id)
        w.delete()
        return Response({"deleted": True})
    except Withdrawal.DoesNotExist:
        return Response({"error": "No encontrado."}, status=404)
    
@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def admin_finance_by_product(request):
    """GET /api/v1/admin/finance/by-product/"""
    

    # Solo órdenes pagadas
    items = OrderItem.objects.filter(
        order__status__in=["paid", "shipped", "completed"]
    ).select_related("product").values(
        "product__name", "product__sku", "product__cost_price"
    ).annotate(
        units_sold    = Sum("quantity"),
        total_revenue = Sum(
            ExpressionWrapper(F("price") * F("quantity"),
            output_field=DecimalField())
        ),
    ).order_by("-total_revenue")

    results = []
    for item in items:
        cost_price    = item["product__cost_price"] or Decimal("0")
        total_cost    = cost_price * item["units_sold"]
        profit        = item["total_revenue"] - total_cost
        margin        = round(float(profit) / float(item["total_revenue"]) * 100, 1) \
                        if item["total_revenue"] else 0

        results.append({
            "product_name":  item["product__name"],
            "sku":           item["product__sku"],
            "units_sold":    item["units_sold"],
            "total_revenue": float(item["total_revenue"]),
            "total_cost":    float(total_cost),
            "profit":        float(profit),
            "margin":        margin,
        })

    return Response({"results": results})    


# ── INVENTARIO INTELIGENTE ────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def inventory_summary(request):
    """KPIs generales del inventario"""
    
    products = Product.objects.filter(is_active=True)

    # Valor total del inventario (stock × cost_price)
    valor_inventario = sum(
        (p.stock * p.cost_price)
        for p in products
        if p.cost_price and p.stock > 0
    )

    # Valor a precio de venta
    valor_venta = sum(
        (p.stock * p.price)
        for p in products
        if p.stock > 0
    )

    total_productos  = products.count()
    sin_stock        = products.filter(stock=0).count()
    stock_critico    = products.filter(stock__gt=0, stock__lte=F("stock_min")).count()
    stock_ok         = products.filter(stock__gt=F("stock_min")).count()

    return Response({
        "valor_inventario":  float(valor_inventario),
        "valor_venta":       float(valor_venta),
        "ganancia_potencial": float(valor_venta - valor_inventario),
        "total_productos":   total_productos,
        "sin_stock":         sin_stock,
        "stock_critico":     stock_critico,
        "stock_ok":          stock_ok,
    })


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def inventory_alerts(request):
    """Productos con stock bajo o sin stock"""

    products = Product.objects.filter(
        is_active=True,
        stock__lte=F("stock_min")
    ).select_related("category", "brand").order_by("stock")

    data = []
    for p in products:
        primary_img = p.images.filter(is_primary=True).first() or p.images.first()
        data.append({
            "id":         str(p.id),
            "name":       p.name,
            "sku":        p.sku,
            "stock":      p.stock,
            "stock_min":  p.stock_min,
            "price":      float(p.price),
            "cost_price": float(p.cost_price) if p.cost_price else None,
            "category":   p.category.name if p.category else None,
            "brand":      p.brand.name if p.brand else None,
            "image":      primary_img.image.url if primary_img else None,
            "status":     "sin_stock" if p.stock == 0 else "critico",
        })

    return Response(data)


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def inventory_rotation(request):
    """Productos más y menos vendidos con rotación"""
    from apps.orders.models import OrderItem
    from apps.products.models import Product
    from django.db.models import Sum, Count, F

    # Ventas por producto (últimos 30 días)
    from django.utils import timezone
    from datetime import timedelta
    hace_30 = timezone.now() - timedelta(days=30)

    ventas = OrderItem.objects.filter(
        order__created_at__gte=hace_30,
        order__status__in=["paid", "shipped", "completed"]
    ).values(
        "product__id", "product__name", "product__stock",
        "product__price", "product__cost_price"
    ).annotate(
        unidades_vendidas=Sum("quantity"),
        ingresos=Sum(F("quantity") * F("price")),
    ).order_by("-unidades_vendidas")[:20]

    data = []
    for v in ventas:
        data.append({
            "product_id":       str(v["product__id"]) if v["product__id"] else None,
            "name":             v["product__name"],
            "stock_actual":     v["product__stock"],
            "unidades_vendidas":v["unidades_vendidas"] or 0,
            "ingresos":         float(v["ingresos"] or 0),
        })

    return Response(data)


@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def inventory_value_by_category(request):
    """Valor del inventario por categoría"""

    categories = Category.objects.filter(is_active=True)
    data = []

    for cat in categories:
        products = Product.objects.filter(
            category=cat, is_active=True, stock__gt=0
        )
        valor = sum(
            float(p.stock * p.cost_price)
            for p in products if p.cost_price
        )
        unidades = sum(p.stock for p in products)
        data.append({
            "category": cat.name,
            "valor_inventario": valor,
            "unidades": unidades,
            "productos": products.count(),
        })

    return Response(sorted(data, key=lambda x: x["valor_inventario"], reverse=True))

@api_view(["GET"])
@permission_classes([IsAdminOrStaff])
def executive_dashboard(request):
    """Dashboard ejecutivo — KPIs del negocio para mobile"""

    hoy        = timezone.now()
    hoy_inicio = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
    ayer_inicio = hoy_inicio - timedelta(days=1)
    semana_inicio = hoy_inicio - timedelta(days=7)
    mes_inicio    = hoy_inicio - timedelta(days=30)
    mes_anterior  = hoy_inicio - timedelta(days=60)

    # ── Ventas ───────────────────────────────────────────────
    def get_ventas(desde, hasta=None):
        qs = Order.objects.filter(
            status__in=["paid", "shipped", "completed"],
            created_at__gte=desde,
        )
        if hasta:
            qs = qs.filter(created_at__lt=hasta)
        return qs.aggregate(
            total=Sum("total"),
            count=Count("id"),
        )

    ventas_hoy        = get_ventas(hoy_inicio)
    ventas_ayer       = get_ventas(ayer_inicio, hoy_inicio)
    ventas_semana     = get_ventas(semana_inicio)
    ventas_mes        = get_ventas(mes_inicio)
    ventas_mes_ant    = get_ventas(mes_anterior, mes_inicio)

    def variacion(actual, anterior):
        a = float(actual or 0)
        p = float(anterior or 0)
        if p == 0:
            return 100 if a > 0 else 0
        return round(((a - p) / p) * 100, 1)

    # ── Órdenes del día ──────────────────────────────────────
    ordenes_hoy = Order.objects.filter(
        created_at__gte=hoy_inicio
    ).select_related().order_by("-created_at")[:5]

    ordenes_hoy_data = [{
        "id":     str(o.id),
        "status": o.status,
        "total":  float(o.total),
        "email":  o.email,
        "hora":   o.created_at.strftime("%H:%M"),
    } for o in ordenes_hoy]

    # ── Inventario crítico ───────────────────────────────────
    
    criticos = Product.objects.filter(
        is_active=True,
        stock__lte=Fdb("stock_min")
    ).count()

    sin_stock = Product.objects.filter(
        is_active=True, stock=0
    ).count()

    # ── Utilidad neta del mes ───────────────────────────────
    try:
        utilidad_cat = DistributionCategory.objects.filter(
            cat_type="remainder"
        ).first()
        utilidad_mes = IncomeDistribution.objects.filter(
            category=utilidad_cat,
            record__order__created_at__gte=mes_inicio,
        ).aggregate(t=Sum("amount"))["t"] or 0
    except Exception:
        utilidad_mes = 0

    # ── Gráfico ventas 7 días ────────────────────────────────
    ventas_7dias = []
    for i in range(6, -1, -1):
        dia_inicio = hoy_inicio - timedelta(days=i)
        dia_fin    = dia_inicio + timedelta(days=1)
        v = Order.objects.filter(
            status__in=["paid", "shipped", "completed"],
            created_at__gte=dia_inicio,
            created_at__lt=dia_fin,
        ).aggregate(total=Sum("total"), count=Count("id"))
        ventas_7dias.append({
            "dia":   dia_inicio.strftime("%a"),
            "fecha": dia_inicio.strftime("%d/%m"),
            "total": float(v["total"] or 0),
            "count": v["count"] or 0,
        })

    # ── Top 3 productos del mes ──────────────────────────────
    from apps.orders.models import OrderItem
    top_productos = OrderItem.objects.filter(
        order__status__in=["paid", "shipped", "completed"],
        order__created_at__gte=mes_inicio,
    ).values("product_name").annotate(
        unidades=Sum("quantity"),
        ingresos=Sum(F("quantity") * F("price")),
    ).order_by("-ingresos")[:3]

    return Response({
        "ventas": {
            "hoy":            float(ventas_hoy["total"] or 0),
            "hoy_count":      ventas_hoy["count"] or 0,
            "ayer":           float(ventas_ayer["total"] or 0),
            "semana":         float(ventas_semana["total"] or 0),
            "mes":            float(ventas_mes["total"] or 0),
            "mes_anterior":   float(ventas_mes_ant["total"] or 0),
            "var_dia":        variacion(ventas_hoy["total"], ventas_ayer["total"]),
            "var_mes":        variacion(ventas_mes["total"], ventas_mes_ant["total"]),
        },
        "utilidad_mes":    float(utilidad_mes),
        "inventario": {
            "criticos":  criticos,
            "sin_stock": sin_stock,
        },
        "ordenes_hoy":     ordenes_hoy_data,
        "ventas_7dias":    ventas_7dias,
        "top_productos":   list(top_productos),
    })